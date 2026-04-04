from fastapi import APIRouter, HTTPException, status, Depends, Query, UploadFile, File
from typing import Optional, Dict, Any
from app.models.device import DeviceCreate, DeviceUpdate, DeviceType
from app.services import device_service, notification_service, defect_service
from app.middleware.auth_middleware import get_current_user, require_admin_or_manager,require_management

router = APIRouter()

MAX_UPLOAD_FILE_SIZE = 10 * 1024 * 1024  # 10MB


def _is_likely_text(content: bytes) -> bool:
    if not content:
        return True
    return b"\x00" not in content


def _validate_upload_signature(filename_lower: str, content: bytes) -> None:
    if filename_lower.endswith(".xlsx"):
        if not content.startswith(b"PK\x03\x04"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLSX file content"
            )
        return

    if filename_lower.endswith(".xls"):
        if not content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLS file content"
            )
        return

    if filename_lower.endswith(".csv"):
        if not _is_likely_text(content[:2048]):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid CSV file content"
            )
        return


@router.get("")
async def get_devices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    device_type: Optional[str] = None,
    holder_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all devices with pagination and filters"""
    try:
        # Filter by holder for non-admin/manager/staff users
        if current_user["role"] not in ["super_admin", "manager", "pdic_staff"]:
            holder_id = current_user["id"]

        result = await device_service.get_devices(
            page=page,
            page_size=page_size,
            status=status,
            device_type=device_type,
            holder_id=holder_id,
            search=search
        )

        return {
            "success": True,
            "message": "Devices retrieved successfully",
            "data": result["data"],
            "pagination": result["pagination"]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve devices: {str(e)}"
        )


@router.get("/for-replacement")
async def get_devices_for_replacement(
    exclude_device_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all devices available as replacements (status=available or returned).
    Management only — returns full stock regardless of holder. Used in the Replace Device modal."""
    if current_user["role"] not in ["super_admin", "manager", "pdic_staff"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only management can access replacement device pool"
        )
    try:
        devices = await device_service.get_devices_for_replacement(
            exclude_device_id=exclude_device_id
        )
        return {
            "success": True,
            "message": "Replacement-eligible devices retrieved successfully",
            "data": devices
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve replacement devices: {str(e)}"
        )


@router.get("/available")
async def get_available_devices(
    current_user: dict = Depends(get_current_user)
):
    """Get devices available to distribute for the current user.
    - admin/manager/staff: PDIC stock (status='available')
    - sub_distributor/cluster/operator: all devices they currently hold"""
    try:
        role = current_user["role"]
        if role in ["super_admin", "manager", "pdic_staff"]:
            devices = await device_service.get_available_devices(holder_id=None)
        else:
            # Sub-level roles can redistribute any device they hold
            devices = await device_service.get_held_devices(holder_id=current_user["id"])

        return {
            "success": True,
            "message": "Available devices retrieved successfully",
            "data": devices
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve available devices: {str(e)}"
        )


@router.get("/my-overview")
async def get_my_device_overview(
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive device overview: devices in hand + under hierarchy + distribution stats.
    - admin/manager/staff: all system devices with full stats
    - sub_distributor: held + cluster + operator devices in their chain
    - cluster: held + operator devices under them
    - operator: only their held devices"""
    try:
        role = current_user["role"]
        if role in ["super_admin", "manager", "pdic_staff"]:
            result = await device_service.get_devices(page=1, page_size=2000)
            all_devices = result["data"]
            stats = await device_service.get_device_stats()
            return {
                "success": True,
                "data": {
                    "held_by_me": all_devices,
                    "under_subordinates": [],
                    "all_under_me": all_devices,
                    "stats": {
                        "in_my_hand": stats.get("total", 0),
                        "under_subordinates": 0,
                        "total_in_chain": stats.get("total", 0),
                        "total_devices_received": 0,
                        "total_devices_sent": 0,
                        "total_distributions_received": 0,
                        "total_distributions_sent": 0,
                        **stats
                    }
                }
            }
        else:
            overview = await device_service.get_user_device_overview(
                user_id=current_user["id"],
                user_role=role
            )
            return {"success": True, "data": overview}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get device overview: {str(e)}"
        )


@router.post("/{device_id}/repair-holder")
async def repair_device_holder(
    device_id: str,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Admin/Manager: repair a device's current_holder by replaying the most recent
    distributed history entry. Use when a double-approval has overwritten the holder."""
    try:
        device = await device_service.repair_device_holder_from_history(device_id)
        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found or no distribution history available"
            )
        return {
            "success": True,
            "message": "Device holder repaired successfully",
            "data": device
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to repair device holder: {str(e)}"
        )



@router.post("/{device_id}/request-edit")
async def request_device_edit(
    device_id: str,
    payload: Dict[str, Any],
    current_user: dict = Depends(get_current_user)
):
    """Staff: request an edit to a device. Sends an approval notification to admins/managers.
    The device is NOT modified until a manager/admin reviews and applies the change."""
    if current_user["role"] not in ["pdic_staff", "super_admin", "manager"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only management users can request device edits"
        )

    try:
        device = await device_service.get_device_by_id(device_id)
        if not device:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Device not found")

        proposer_name = current_user.get("name") or current_user.get("email", "pdic_staff")
        changes_summary = ", ".join(
            f"{k}: '{v}'" for k, v in payload.items()
            if k not in ("_edit_note",) and v
        )
        message = (
            f"Staff Edit Request from {proposer_name}:\n"
            f"Device: {device.get('device_id')} (Serial: {device.get('serial_number')})\n"
            f"Proposed Changes: {changes_summary or 'No changes specified'}"
        )

        from app.database import get_db, rows_to_list
        async with get_db() as db:
            cursor = await db.execute(
                "SELECT id FROM users WHERE role IN ('super_admin', 'manager')"
            )
            rows = await cursor.fetchall()
            admin_ids = [str(row[0]) for row in rows]

        for admin_id in admin_ids:
            await notification_service.create_notification(
                user_id=admin_id,
                title="⚙️ Device Edit Approval Request",
                message=message,
                notification_type="device_edit_request",
                reference_id=device_id,
                reference_type="device"
            )

        return {
            "success": True,
            "message": f"Edit request sent to {len(admin_ids)} manager(s)/admin(s) for approval."
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to submit edit request: {str(e)}"
        )


@router.get("/track/{serial_number}")
async def track_device_by_serial(
    serial_number: str,
    current_user: dict = Depends(get_current_user)
):
    """Track device by serial number with full history"""
    try:
        device = await device_service.track_device_by_serial(serial_number)

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        return {
            "success": True,
            "message": "Device tracked successfully",
            "data": device
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to track device '{serial_number}': {str(e)}"
        )


@router.get("/{device_id}")
async def get_device(
    device_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get device by ID"""
    try:
        device = await device_service.get_device_by_id(device_id)

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        return {
            "success": True,
            "message": "Device retrieved successfully",
            "data": device
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve device '{device_id}': {str(e)}"
        )


@router.get("/{device_id}/history")
async def get_device_history(
    device_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get device history"""
    try:
        device = await device_service.get_device_by_id(device_id)

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        history = await device_service.get_device_history(device_id)

        return {
            "success": True,
            "message": "Device history retrieved successfully",
            "data": history
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve history for device '{device_id}': {str(e)}"
        )


@router.post("/bulk-upload", status_code=status.HTTP_201_CREATED)
async def bulk_upload_devices(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_management)
):
    """Bulk upload devices from an Excel file.

    Supported schemas (case-insensitive headers):
    - SB sheet: vendor, device_type, model, nuid, box_type
    - Regular sheet: vendor, device_type, model, mac_address, serial_number, band_type

    Alias support: manufacturer -> vendor, SB/set-top box/set top box/stb -> Set-top box
    """
    filename_lower = file.filename.lower()
    if not filename_lower.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel (.xlsx, .xls) or CSV (.csv) files are supported"
        )

    try:
        import io

        contents = await file.read()
        if len(contents) > MAX_UPLOAD_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="File too large. Maximum 10MB allowed"
            )

        _validate_upload_signature(filename_lower, contents)

        if filename_lower.endswith('.csv'):
            import csv
            decoded = contents.decode('utf-8-sig')  # strip BOM if present
            reader = csv.reader(io.StringIO(decoded))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV file is empty")
            headers = [h.strip().lower() for h in all_rows[0]]
            data_rows = all_rows[1:]

            def iter_data_rows():
                for row in data_rows:
                    # Pad short rows to header length
                    padded = row + [''] * (len(headers) - len(row))
                    yield tuple(padded[:len(headers)])
        else:
            if filename_lower.endswith('.xls'):
                import xlrd
                wb = xlrd.open_workbook(file_contents=contents)
                ws = wb.sheet_by_index(0)
                headers = [str(ws.cell_value(0, col)).strip().lower() for col in range(ws.ncols)]

                def iter_data_rows():
                    for row_idx in range(1, ws.nrows):
                        yield tuple(ws.cell_value(row_idx, col) for col in range(ws.ncols))
            else:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                def iter_data_rows():
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        yield row

        normalized_headers = ["vendor" if h == "manufacturer" else h for h in headers]
        header_set = set(normalized_headers)

        sb_required = {"vendor", "device_type", "model", "nuid", "box_type"}
        regular_required = {"vendor", "device_type", "model", "mac_address", "serial_number"}

        has_sb_schema = sb_required.issubset(header_set)
        has_regular_schema = regular_required.issubset(header_set)
        if not has_sb_schema and not has_regular_schema:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Missing required columns. Expected either SB schema "
                    "(vendor, device_type, model, nuid, box_type) or regular schema "
                    "(vendor, device_type, model, mac_address, serial_number)."
                )
            )

        valid_types = {t.value.lower(): t.value for t in DeviceType}
        valid_types.update({
            "sb": DeviceType.SETUP_BOX.value,
            "set top box": DeviceType.SETUP_BOX.value,
            "stb": DeviceType.SETUP_BOX.value,
        })
        valid_bands = {
            "single_band": "single_band",
            "single band": "single_band",
            "single": "single_band",
            "dual_band": "dual_band",
            "dual band": "dual_band",
            "dual": "dual_band",
        }
        created, skipped, errors = [], [], []

        for row_idx, row in enumerate(iter_data_rows(), start=2):
            if row is None:
                row = ()

            row_data = {
                normalized_headers[i]: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
                for i in range(len(normalized_headers))
            }

            # Skip completely empty rows
            if not any(row_data.values()):
                continue

            # Normalise device_type
            raw_type = row_data.get("device_type", "").lower()
            device_type_val = valid_types.get(raw_type)
            if not device_type_val:
                errors.append({"row": row_idx, "error": f"Invalid device_type '{row_data.get('device_type')}'"})
                continue

            is_sb_row = device_type_val == DeviceType.SETUP_BOX.value
            vendor = row_data.get("vendor", "")
            model = row_data.get("model", "")
            nuid = row_data.get("nuid", "")
            serial = row_data.get("serial_number", "")
            mac = row_data.get("mac_address", "")

            if not vendor:
                errors.append({"row": row_idx, "serial": serial or nuid or "", "error": "Missing vendor"})
                continue
            if not model:
                errors.append({"row": row_idx, "serial": serial or nuid or "", "error": "Missing model"})
                continue

            if is_sb_row:
                if not nuid:
                    errors.append({"row": row_idx, "error": "Missing nuid for SB row"})
                    continue
                box_type = str(row_data.get("box_type", "")).strip().upper()
                if not box_type:
                    errors.append({"row": row_idx, "error": "Missing box_type for SB row"})
                    continue
                if box_type not in {"HD", "OTT"}:
                    errors.append({"row": row_idx, "error": "Invalid box_type. Use HD or OTT"})
                    continue
                band_type_val = None
            else:
                if not serial:
                    errors.append({"row": row_idx, "error": "Missing serial_number"})
                    continue
                if not mac:
                    errors.append({"row": row_idx, "serial": serial, "error": "Missing mac_address"})
                    continue
                raw_band = row_data.get("band_type", "").lower()
                band_type_val = valid_bands.get(raw_band) if raw_band else None
                if raw_band and not band_type_val:
                    errors.append({"row": row_idx, "serial": serial, "error": f"Invalid band_type '{row_data.get('band_type')}'"})
                    continue

            try:
                device_data = DeviceCreate(
                    device_type=device_type_val,
                    model=model,
                    serial_number=None if is_sb_row else serial,
                    mac_address=None if is_sb_row else mac,
                    manufacturer=vendor,
                    band_type=band_type_val,
                    box_type=box_type if is_sb_row else None,
                    nuid=nuid or None,
                    metadata=(
                        {"box_type": box_type}
                        if is_sb_row and box_type
                        else None
                    ),
                )
                device = await device_service.create_device(
                    device_data=device_data,
                    created_by=current_user["id"],
                    created_by_name=current_user["name"]
                )
                if not device:
                    errors.append({"row": row_idx, "serial": serial or nuid, "error": "Device was not created"})
                    continue
                created.append(device["device_id"])
            except ValueError as e:
                skipped.append({"row": row_idx, "serial": serial or nuid, "reason": str(e)})
            except Exception as e:
                errors.append({"row": row_idx, "serial": serial or nuid, "error": str(e)})

        return {
            "success": True,
            "message": f"Bulk upload complete: {len(created)} created, {len(skipped)} skipped, {len(errors)} errors",
            "data": {
                "created_count": len(created),
                "skipped_count": len(skipped),
                "error_count": len(errors),
                "created": created,
                "skipped": skipped,
                "errors": errors,
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process file: {str(e)}"
        )


@router.post("", status_code=status.HTTP_201_CREATED)
async def create_device(
    device_data: DeviceCreate,
    current_user: dict = Depends(require_management)
):
    """Register a new device"""
    try:
        device = await device_service.create_device(
            device_data=device_data,
            created_by=current_user["id"],
            created_by_name=current_user["name"]
        )

        return {
            "success": True,
            "message": "Device registered successfully",
            "data": device
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to register device: {str(e)}"
        )


@router.put("/{device_id}")
async def update_device(
    device_id: str,
    device_data: DeviceUpdate,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Update device"""
    try:
        device = await device_service.update_device(device_id, device_data)

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        return {
            "success": True,
            "message": "Device updated successfully",
            "data": device
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update device '{device_id}': {str(e)}"
        )


@router.delete("/{device_id}")
async def delete_device(
    device_id: str,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Delete device"""
    try:
        success = await device_service.delete_device(device_id)

        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        return {
            "success": True,
            "message": "Device deleted successfully"
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete device '{device_id}': {str(e)}"
        )


@router.patch("/{device_id}/status")
async def update_device_status(
    device_id: str,
    status_update: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update device status"""
    status_value = status_update.get("status")
    notes = status_update.get("notes")
    
    if not status_value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Status is required"
        )

    try:
        device = await device_service.update_device_status(
            device_id=device_id,
            status=status_value,
            performed_by=current_user["id"],
            performed_by_name=current_user["name"],
            notes=notes
        )

        if str(status_value).lower() == "defective":
            await defect_service.create_or_get_active_defect_for_device(
                device_id=device_id,
                reporter=current_user,
                notes=notes
            )

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        return {
            "success": True,
            "message": "Device status updated successfully",
            "data": device
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update status for device '{device_id}': {str(e)}"
        )

