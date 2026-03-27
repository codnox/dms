import csv
import io

from fastapi import APIRouter, HTTPException, status, Depends, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
from typing import Optional
from pydantic import BaseModel
from app.models.distribution import DistributionCreate, DistributionStatusUpdate
from app.services import distribution_service
from app.middleware.auth_middleware import get_current_user, require_admin_or_manager, require_management

router = APIRouter()


class ReceiptConfirmation(BaseModel):
    received: bool
    notes: Optional[str] = None


@router.post("/bulk-upload")
async def bulk_upload_distribution(
    file: UploadFile = File(...),
    to_user_id: str = Form(...),
    notes: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Create a distribution from uploaded CSV/Excel rows using mac_address and/or nuid."""
    filename_lower = (file.filename or "").lower()
    if not filename_lower.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel (.xlsx, .xls) or CSV (.csv) files are supported"
        )

    try:
        contents = await file.read()

        if filename_lower.endswith(".csv"):
            decoded = contents.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(decoded))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="CSV file is empty"
                )
            headers = [str(h).strip().lower() for h in all_rows[0]]
            data_rows = all_rows[1:]

            def iter_data_rows():
                for row in data_rows:
                    padded = row + [""] * (len(headers) - len(row))
                    yield tuple(padded[:len(headers)])

        else:
            import openpyxl

            workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            worksheet = workbook.active
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1), None)
            if not header_row:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Excel file is empty"
                )

            headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in header_row]

            def iter_data_rows():
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    yield row

        if "mac_address" not in headers and "nuid" not in headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Missing required columns: add at least one of mac_address or nuid"
            )

        identifier_rows = []
        for row_idx, row in enumerate(iter_data_rows(), start=2):
            row_data = {
                headers[i]: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
                for i in range(len(headers))
            }

            mac_address = row_data.get("mac_address", "")
            nuid = row_data.get("nuid", "")

            if not mac_address and not nuid:
                # Skip fully empty lines, otherwise keep for validation.
                if not any(v for v in row_data.values()):
                    continue

            identifier_rows.append({
                "row": row_idx,
                "mac_address": mac_address,
                "nuid": nuid,
            })

        if not identifier_rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No identifier rows found in file"
            )

        result = await distribution_service.create_distribution_from_identifiers(
            to_user_id=to_user_id,
            identifier_rows=identifier_rows,
            from_user=current_user,
            notes=notes,
        )

        if result["created"]:
            message = (
                f"Distribution created successfully with {result['created_count']} device(s)"
            )
        else:
            message = (
                f"Upload validation failed: {result['error_count']} row error(s). "
                "Fix errors and upload again."
            )

        return {
            "success": True,
            "message": message,
            "data": result,
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process file: {str(e)}"
        )


@router.post("/sync-devices")
async def sync_distribution_devices(
    current_user: dict = Depends(require_admin_or_manager)
):
    """Sync device holders for all approved distributions (admin fix endpoint)"""
    try:
        result = await distribution_service.sync_approved_distributions(user=current_user)
        return {
            "success": True,
            "message": f"Synced {result['devices_synced']} device(s) from {result['total_distributions']} approved distribution(s)",
            "data": result
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@router.get("")
async def get_distributions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all distributions with pagination and filters"""
    try:
        # Filter by user for non-admin/manager/staff
        user_id = None
        if current_user["role"] not in ["admin", "manager", "staff"]:
            user_id = current_user["id"]

        result = await distribution_service.get_distributions(
            page=page,
            page_size=page_size,
            status=status,
            user_id=user_id,
            search=search
        )

        return {
            "success": True,
            "message": "Distributions retrieved successfully",
            "data": result["data"],
            "pagination": result["pagination"]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve distributions: {str(e)}"
        )


@router.get("/pending")
async def get_pending_distributions(
    current_user: dict = Depends(require_management)
):
    """Get pending distributions for approval"""
    try:
        distributions = await distribution_service.get_pending_distributions()

        return {
            "success": True,
            "message": "Pending distributions retrieved successfully",
            "data": distributions
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve pending distributions: {str(e)}"
        )


@router.get("/{distribution_id}/manifest")
async def download_distribution_manifest(
    distribution_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download generated Excel manifest for a distribution."""
    try:
        manifest = await distribution_service.get_distribution_manifest_file(distribution_id, current_user)
        if not manifest:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Distribution manifest not found"
            )

        return FileResponse(
            path=manifest["path"],
            filename=manifest["filename"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to download manifest for distribution '{distribution_id}': {str(e)}"
        )


@router.get("/{distribution_id}")
async def get_distribution(
    distribution_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get distribution by ID"""
    try:
        distribution = await distribution_service.get_distribution_by_id(distribution_id)

        if not distribution:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Distribution not found"
            )

        return {
            "success": True,
            "message": "Distribution retrieved successfully",
            "data": distribution
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve distribution '{distribution_id}': {str(e)}"
        )


@router.post("", status_code=status.HTTP_201_CREATED)
async def create_distribution(
    dist_data: DistributionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new distribution request.
    - admin/manager/staff: can distribute PDIC devices to any sub-level user
    - sub_distributor: can distribute their held devices to clusters or operators under them
    - cluster: can distribute their held devices to operators under them
    - operator: can distribute their held devices to operators in the same cluster
    """
    
    try:
        distribution = await distribution_service.create_distribution(
            dist_data=dist_data,
            from_user=current_user
        )

        return {
            "success": True,
            "message": "Distribution created successfully",
            "data": distribution
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create distribution: {str(e)}"
        )


@router.post("/{distribution_id}/receipt")
async def confirm_distribution_receipt(
    distribution_id: str,
    body: ReceiptConfirmation,
    current_user: dict = Depends(get_current_user)
):
    """Recipient confirms or disputes receipt of a distribution.
    - received=true  → Distribution becomes APPROVED; receiver can now redistribute devices
    - received=false → Distribution becomes DISPUTED; admin/manager + sender are notified
    """
    try:
        distribution = await distribution_service.confirm_receipt(
            distribution_id=distribution_id,
            received=body.received,
            user=current_user,
            notes=body.notes
        )
        action = "confirmed" if body.received else "disputed"
        return {
            "success": True,
            "message": f"Receipt {action} successfully",
            "data": distribution
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to confirm receipt: {str(e)}"
        )


@router.patch("/{distribution_id}/status")
async def update_distribution_status(
    distribution_id: str,
    status_update: DistributionStatusUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update distribution status"""
    try:
        distribution = await distribution_service.update_distribution_status(
            distribution_id=distribution_id,
            status=status_update.status.value,
            user=current_user,
            notes=status_update.notes
        )

        if not distribution:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Distribution not found"
            )

        return {
            "success": True,
            "message": "Distribution status updated successfully",
            "data": distribution
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except PermissionError as e:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update distribution status '{distribution_id}': {str(e)}"
        )


@router.delete("/{distribution_id}")
async def cancel_distribution(
    distribution_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Cancel a distribution (only by creator)"""
    try:
        success = await distribution_service.cancel_distribution(
            distribution_id=distribution_id,
            user=current_user
        )

        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Distribution not found"
            )

        return {
            "success": True,
            "message": "Distribution cancelled successfully"
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to cancel distribution '{distribution_id}': {str(e)}"
        )
