from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any
import csv
import io

from openpyxl import Workbook

from app.database import get_db, rows_to_list


ALLOWED_REPORT_TABLES = {
    "devices",
    "distributions",
    "defects",
    "returns",
    "users",
}


async def _count(db, table: str, condition: str = "1=1", params=()) -> int:
    if table not in ALLOWED_REPORT_TABLES:
        raise ValueError(f"Invalid table name: {table}")
    cursor = await db.execute(f"SELECT COUNT(*) FROM {table} WHERE {condition}", params)
    return (await cursor.fetchone())[0]


async def get_inventory_report() -> Dict[str, Any]:
    """Generate device inventory report"""
    async with get_db() as db:
        total = await _count(db, "devices")

        by_status = {}
        for status in ["available", "distributed", "in_use", "defective", "returned", "maintenance"]:
            by_status[status] = await _count(db, "devices", "status = ?", (status,))

        # By type
        cursor = await db.execute("SELECT DISTINCT device_type FROM devices")
        dtypes = [r[0] for r in await cursor.fetchall()]
        by_type = {}
        for dtype in dtypes:
            by_type[dtype] = await _count(db, "devices", "device_type = ?", (dtype,))

        # By holder type
        cursor = await db.execute("SELECT DISTINCT current_holder_type FROM devices WHERE current_holder_type IS NOT NULL")
        htypes = [r[0] for r in await cursor.fetchall()]
        by_location = {}
        for htype in htypes:
            by_location[htype] = await _count(db, "devices", "current_holder_type = ?", (htype,))

        return {
            "total_devices": total,
            "by_status": by_status,
            "by_type": by_type,
            "by_location": by_location,
            "generated_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        }


async def get_distribution_summary() -> Dict[str, Any]:
    """Generate distribution summary report"""
    async with get_db() as db:
        total = await _count(db, "distributions")

        by_status = {}
        for status in ["pending", "approved", "in_transit", "delivered", "rejected", "cancelled"]:
            by_status[status] = await _count(db, "distributions", "status = ?", (status,))

        by_month = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        for i in range(5, -1, -1):
            month_start = datetime(now.year, now.month, 1) - timedelta(days=i * 30)
            month_end = month_start + timedelta(days=30)
            count = await _count(db, "distributions", "created_at >= ? AND created_at < ?",
                                 (month_start.isoformat(), month_end.isoformat()))
            by_month.append({"month": month_start.strftime("%B %Y"), "count": count})

        # Top distributors
        cursor = await db.execute(
            """SELECT to_user_name, SUM(device_count) as total
            FROM distributions WHERE status = 'delivered'
            GROUP BY to_user_name ORDER BY total DESC LIMIT 5"""
        )
        top = await cursor.fetchall()

        return {
            "total": total,
            "by_status": by_status,
            "by_month": by_month,
            "top_distributors": [{"name": r[0], "devices": r[1]} for r in top],
            "generated_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        }


async def get_defect_summary() -> Dict[str, Any]:
    """Generate defect summary report"""
    async with get_db() as db:
        total = await _count(db, "defects")

        by_status = {}
        for status in ["reported", "under_review", "approved", "rejected", "resolved"]:
            by_status[status] = await _count(db, "defects", "status = ?", (status,))

        by_severity = {}
        for severity in ["critical", "high", "medium", "low"]:
            by_severity[severity] = await _count(db, "defects", "severity = ?", (severity,))

        by_type = {}
        for defect_type in ["hardware", "software", "physical_damage", "performance", "connectivity", "other"]:
            by_type[defect_type] = await _count(db, "defects", "defect_type = ?", (defect_type,))

        by_month = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        for i in range(5, -1, -1):
            month_start = datetime(now.year, now.month, 1) - timedelta(days=i * 30)
            month_end = month_start + timedelta(days=30)
            count = await _count(db, "defects", "created_at >= ? AND created_at < ?",
                                 (month_start.isoformat(), month_end.isoformat()))
            by_month.append({"month": month_start.strftime("%B %Y"), "count": count})

        return {
            "total": total,
            "by_status": by_status,
            "by_severity": by_severity,
            "by_type": by_type,
            "by_month": by_month,
            "generated_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        }


async def get_return_summary() -> Dict[str, Any]:
    """Generate return summary report"""
    async with get_db() as db:
        total = await _count(db, "returns")

        by_status = {}
        for status in ["pending", "approved", "in_transit", "received", "rejected", "cancelled"]:
            by_status[status] = await _count(db, "returns", "status = ?", (status,))

        by_reason = {}
        for reason in ["defective", "unused", "end_of_contract", "upgrade", "other"]:
            by_reason[reason] = await _count(db, "returns", "reason = ?", (reason,))

        by_month = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        for i in range(5, -1, -1):
            month_start = datetime(now.year, now.month, 1) - timedelta(days=i * 30)
            month_end = month_start + timedelta(days=30)
            count = await _count(db, "returns", "created_at >= ? AND created_at < ?",
                                 (month_start.isoformat(), month_end.isoformat()))
            by_month.append({"month": month_start.strftime("%B %Y"), "count": count})

        return {
            "total": total,
            "by_status": by_status,
            "by_reason": by_reason,
            "by_month": by_month,
            "generated_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        }


async def get_user_activity_report() -> Dict[str, Any]:
    """Generate user activity report"""
    async with get_db() as db:
        by_role = {}
        for role in ["super_admin", "manager", "pdic_staff", "sub_distributor", "cluster", "operator"]:
            by_role[role] = await _count(db, "users", "role = ?", (role,))

        thirty_days_ago = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=30)).isoformat()
        active_users = await _count(db, "users", "last_login >= ?", (thirty_days_ago,))
        total_users = await _count(db, "users")

        cursor = await db.execute("SELECT * FROM device_history ORDER BY timestamp DESC LIMIT 50")
        rows = await cursor.fetchall()

        return {
            "total_users": total_users,
            "active_users": active_users,
            "by_role": by_role,
            "recent_activities": rows_to_list(rows),
            "generated_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        }


async def get_device_utilization_report() -> Dict[str, Any]:
    """Generate device utilization report"""
    async with get_db() as db:
        total_devices = await _count(db, "devices")
        in_use = await _count(db, "devices", "status IN ('distributed', 'in_use')")
        available = await _count(db, "devices", "status = 'available'")
        defective = await _count(db, "devices", "status = 'defective'")

        utilization_rate = (in_use / total_devices * 100) if total_devices > 0 else 0

        return {
            "total_devices": total_devices,
            "in_use": in_use,
            "available": available,
            "defective": defective,
            "utilization_rate": round(utilization_rate, 2),
            "generated_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        }


def _build_device_journey(history_rows: List[Dict[str, Any]], device_row: Dict[str, Any]) -> Dict[str, str]:
    """Build start, intermediate path, and current location strings for a device."""
    sorted_history = sorted(history_rows, key=lambda row: str(row.get("timestamp") or ""))

    start_location = "PDIC"
    for row in sorted_history:
        action = str(row.get("action") or "").lower()
        location = str(row.get("location") or "").strip()
        if action == "registered" and location:
            start_location = location
            break

    path_nodes = [start_location]
    for row in sorted_history:
        action = str(row.get("action") or "").lower()
        if action != "distributed":
            continue

        next_point = str(row.get("to_user_name") or "").strip() or str(row.get("location") or "").strip()
        if next_point and next_point != path_nodes[-1]:
            path_nodes.append(next_point)

    current_location = (
        str(device_row.get("current_holder_name") or "").strip()
        or str(device_row.get("current_location") or "").strip()
        or (path_nodes[-1] if path_nodes else "")
    )

    full_path_nodes = path_nodes[:]
    if current_location and (not full_path_nodes or current_location != full_path_nodes[-1]):
        full_path_nodes.append(current_location)

    passed_through_nodes = []
    if len(full_path_nodes) >= 3:
        passed_through_nodes = full_path_nodes[1:-1]

    return {
        "started_from": start_location,
        "passed_through": " -> ".join(passed_through_nodes),
        "current_at": current_location,
        "journey_path": " -> ".join(full_path_nodes),
    }


def _build_device_backup_file(rows: List[Dict[str, Any]], file_format: str) -> Dict[str, Any]:
    """Build downloadable backup payload for devices in CSV or XLSX format."""
    headers = [
        "device_db_id",
        "device_id",
        "serial_number",
        "mac_address",
        "nuid",
        "device_type",
        "model",
        "manufacturer",
        "status",
        "current_holder_name",
        "current_holder_type",
        "started_from",
        "passed_through",
        "current_at",
        "journey_path",
        "created_at",
        "updated_at",
    ]

    normalized = str(file_format or "xlsx").strip().lower()
    generated_ts = datetime.now(timezone.utc).replace(tzinfo=None).strftime("%Y%m%d_%H%M%S")

    if normalized == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Device Backup"
        sheet.append(headers)

        for row in rows:
            sheet.append([row.get(col, "") for col in headers])

        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return {
            "content": payload.getvalue(),
            "filename": f"device-backup-{generated_ts}.xlsx",
            "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    if normalized == "csv":
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in headers})

        return {
            "content": buffer.getvalue().encode("utf-8"),
            "filename": f"device-backup-{generated_ts}.csv",
            "media_type": "text/csv",
        }

    raise ValueError("Unsupported export format. Use 'csv' or 'xlsx'")


async def get_device_backup_export(file_format: str = "xlsx") -> Dict[str, Any]:
    """Generate a full device backup export including journey path details."""
    async with get_db() as db:
        cursor = await db.execute("SELECT * FROM devices ORDER BY id ASC")
        device_rows = rows_to_list(await cursor.fetchall())

        cursor = await db.execute("SELECT * FROM device_history ORDER BY timestamp ASC")
        history_rows = rows_to_list(await cursor.fetchall())

    history_by_device: Dict[str, List[Dict[str, Any]]] = {}
    for row in history_rows:
        key = str(row.get("device_id") or "")
        if not key:
            continue
        history_by_device.setdefault(key, []).append(row)

    export_rows: List[Dict[str, Any]] = []
    for device in device_rows:
        device_key = str(device.get("id") or "")
        journey = _build_device_journey(history_by_device.get(device_key, []), device)

        export_rows.append(
            {
                "device_db_id": str(device.get("id") or ""),
                "device_id": str(device.get("device_id") or ""),
                "serial_number": str(device.get("serial_number") or ""),
                "mac_address": str(device.get("mac_address") or ""),
                "nuid": str(device.get("nuid") or ""),
                "device_type": str(device.get("device_type") or ""),
                "model": str(device.get("model") or ""),
                "manufacturer": str(device.get("manufacturer") or ""),
                "status": str(device.get("status") or ""),
                "current_holder_name": str(device.get("current_holder_name") or ""),
                "current_holder_type": str(device.get("current_holder_type") or ""),
                "started_from": journey["started_from"],
                "passed_through": journey["passed_through"],
                "current_at": journey["current_at"],
                "journey_path": journey["journey_path"],
                "created_at": str(device.get("created_at") or ""),
                "updated_at": str(device.get("updated_at") or ""),
            }
        )

    return _build_device_backup_file(export_rows, file_format=file_format)


def _build_returns_defects_backup_file(
    returns_rows: List[Dict[str, Any]],
    defects_rows: List[Dict[str, Any]],
    file_format: str,
) -> Dict[str, Any]:
    """Build downloadable backup payload for returns and defects."""
    def _xlsx_cell_value(value: Any) -> Any:
        if value is None:
            return ""
        return value

    normalized = str(file_format or "xlsx").strip().lower()
    generated_ts = datetime.now(timezone.utc).replace(tzinfo=None).strftime("%Y%m%d_%H%M%S")

    return_headers = [
        "return_id",
        "device_identifier",
        "device_model",
        "device_serial",
        "device_type",
        "requested_by_name",
        "return_to_name",
        "reason",
        "description",
        "status",
        "request_date",
        "approval_date",
        "received_date",
        "approved_by_name",
        "created_at",
        "updated_at",
    ]

    defect_headers = [
        "report_id",
        "device_identifier",
        "device_model",
        "device_serial",
        "device_type",
        "reported_by_name",
        "operator_name",
        "sub_distributor_name",
        "defect_type",
        "severity",
        "description",
        "symptoms",
        "report_target",
        "forwarded_to_management",
        "status",
        "resolution",
        "resolved_by_name",
        "resolved_at",
        "created_at",
        "updated_at",
    ]

    if normalized == "xlsx":
        workbook = Workbook()

        returns_sheet = workbook.active
        returns_sheet.title = "Returned Devices"
        returns_sheet.append(return_headers)
        for row in returns_rows:
            returns_sheet.append([_xlsx_cell_value(row.get(col)) for col in return_headers])

        defects_sheet = workbook.create_sheet("Defect Reports")
        defects_sheet.append(defect_headers)
        for row in defects_rows:
            defects_sheet.append([_xlsx_cell_value(row.get(col)) for col in defect_headers])

        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return {
            "content": payload.getvalue(),
            "filename": f"returns-defects-backup-{generated_ts}.xlsx",
            "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    if normalized == "csv":
        combined_headers = [
            "record_type",
            "id",
            "device_identifier",
            "device_model",
            "device_serial",
            "device_type",
            "person_name",
            "category",
            "status",
            "description",
            "created_at",
            "updated_at",
        ]

        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=combined_headers)
        writer.writeheader()

        for row in returns_rows:
            writer.writerow(
                {
                    "record_type": "return",
                    "id": str(row.get("return_id") or ""),
                    "device_identifier": str(row.get("device_identifier") or ""),
                    "device_model": str(row.get("device_model") or ""),
                    "device_serial": str(row.get("device_serial") or ""),
                    "device_type": str(row.get("device_type") or ""),
                    "person_name": str(row.get("requested_by_name") or ""),
                    "category": str(row.get("reason") or ""),
                    "status": str(row.get("status") or ""),
                    "description": str(row.get("description") or ""),
                    "created_at": str(row.get("created_at") or ""),
                    "updated_at": str(row.get("updated_at") or ""),
                }
            )

        for row in defects_rows:
            writer.writerow(
                {
                    "record_type": "defect",
                    "id": str(row.get("report_id") or ""),
                    "device_identifier": str(row.get("device_identifier") or ""),
                    "device_model": str(row.get("device_model") or ""),
                    "device_serial": str(row.get("device_serial") or ""),
                    "device_type": str(row.get("device_type") or ""),
                    "person_name": str(row.get("reported_by_name") or ""),
                    "category": str(row.get("defect_type") or ""),
                    "status": str(row.get("status") or ""),
                    "description": str(row.get("description") or ""),
                    "created_at": str(row.get("created_at") or ""),
                    "updated_at": str(row.get("updated_at") or ""),
                }
            )

        return {
            "content": buffer.getvalue().encode("utf-8"),
            "filename": f"returns-defects-backup-{generated_ts}.csv",
            "media_type": "text/csv",
        }

    raise ValueError("Unsupported export format. Use 'csv' or 'xlsx'")


async def get_returns_defects_backup_export(file_format: str = "xlsx") -> Dict[str, Any]:
    """Generate backup export for return requests and defect reports."""
    async with get_db() as db:
        returns_cursor = await db.execute("SELECT * FROM returns ORDER BY created_at DESC")
        returns_rows = rows_to_list(await returns_cursor.fetchall())

        defects_cursor = await db.execute("SELECT * FROM defects ORDER BY created_at DESC")
        defects_rows = rows_to_list(await defects_cursor.fetchall())

        devices_cursor = await db.execute("SELECT id, device_id, model FROM devices")
        devices_rows = rows_to_list(await devices_cursor.fetchall())

        users_cursor = await db.execute("SELECT id, name FROM users")
        users_rows = rows_to_list(await users_cursor.fetchall())

    device_lookup: Dict[str, Dict[str, Any]] = {}
    for device in devices_rows:
        db_id = str(device.get("id") or "").strip()
        business_id = str(device.get("device_id") or "").strip()
        if db_id:
            device_lookup[db_id] = device
        if business_id:
            device_lookup[business_id] = device

    user_name_lookup: Dict[str, str] = {
        str(user.get("id") or "").strip(): str(user.get("name") or "").strip()
        for user in users_rows
        if str(user.get("id") or "").strip()
    }

    for row in returns_rows:
        raw_device_id = str(row.get("device_id") or "").strip()
        resolved_device = device_lookup.get(raw_device_id)
        row["device_identifier"] = str(
            (resolved_device or {}).get("device_id")
            or raw_device_id
        )
        row["device_model"] = str((resolved_device or {}).get("model") or "")

        if not str(row.get("requested_by_name") or "").strip():
            row["requested_by_name"] = user_name_lookup.get(str(row.get("requested_by") or "").strip(), "")
        if not str(row.get("return_to_name") or "").strip():
            row["return_to_name"] = user_name_lookup.get(str(row.get("return_to") or "").strip(), "")
        if not str(row.get("approved_by_name") or "").strip():
            row["approved_by_name"] = user_name_lookup.get(str(row.get("approved_by") or "").strip(), "")

    for row in defects_rows:
        raw_device_id = str(row.get("device_id") or "").strip()
        resolved_device = device_lookup.get(raw_device_id)
        row["device_identifier"] = str(
            (resolved_device or {}).get("device_id")
            or raw_device_id
        )
        row["device_model"] = str((resolved_device or {}).get("model") or "")

        if not str(row.get("reported_by_name") or "").strip():
            row["reported_by_name"] = user_name_lookup.get(str(row.get("reported_by") or "").strip(), "")
        if not str(row.get("resolved_by_name") or "").strip():
            row["resolved_by_name"] = user_name_lookup.get(str(row.get("resolved_by") or "").strip(), "")

        row["operator_name"] = user_name_lookup.get(str(row.get("operator_id") or "").strip(), "")
        row["sub_distributor_name"] = user_name_lookup.get(str(row.get("sub_distributor_id") or "").strip(), "")

    return _build_returns_defects_backup_file(
        returns_rows=returns_rows,
        defects_rows=defects_rows,
        file_format=file_format,
    )

